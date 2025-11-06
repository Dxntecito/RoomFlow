from bd import get_connection

def get_todas_estadisticas():
    """
    Obtener todas las estadísticas en una sola conexión para mejorar el rendimiento
    Retorna un diccionario con todas las estadísticas
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            # Obtener totales básicos
            cursor.execute("SELECT COUNT(*) FROM CLIENTE")
            total_clientes = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM HABITACION WHERE estado = 1")
            total_habitaciones = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM RESERVA")
            total_reservas = cursor.fetchone()[0]
            
            # Reservas por estado
            cursor.execute("""
                SELECT 
                    CASE 
                        WHEN estado = 1 THEN 'Activa'
                        WHEN estado = 0 THEN 'Cancelada'
                        ELSE 'Desconocida'
                    END as estado_reserva,
                    COUNT(*) as cantidad
                FROM RESERVA
                GROUP BY estado
            """)
            resultados_reservas = cursor.fetchall()
            reservas_por_estado = {}
            for row in resultados_reservas:
                reservas_por_estado[row[0]] = row[1]
            
            # Clientes por tipo
            cursor.execute("""
                SELECT 
                    tc.descripcion as tipo_cliente,
                    COUNT(*) as cantidad
                FROM CLIENTE c
                INNER JOIN TIPO_CLIENTE tc ON c.id_tipo_cliente = tc.tipo_cliente_id
                GROUP BY tc.descripcion
            """)
            resultados_clientes = cursor.fetchall()
            clientes_por_tipo = {}
            for row in resultados_clientes:
                clientes_por_tipo[row[0]] = row[1]
            
            # Empleados por turno
            cursor.execute("""
                SELECT 
                    t.nombre_turno,
                    COUNT(DISTINCT dt.empleado_id) as cantidad
                FROM TURNO t
                LEFT JOIN DETALLE_TURNO dt ON t.turno_id = dt.turno_id
                GROUP BY t.turno_id, t.nombre_turno
            """)
            resultados_turnos = cursor.fetchall()
            empleados_por_turno = {}
            for row in resultados_turnos:
                empleados_por_turno[row[0]] = row[1]
            
            # Total tablas
            cursor.execute("""
                SELECT COUNT(*) 
                FROM information_schema.tables 
                WHERE table_schema = DATABASE()
            """)
            total_tablas = cursor.fetchone()[0]
            
            # Total registros
            cursor.execute("""
                SELECT 
                    (SELECT COUNT(*) FROM CLIENTE) +
                    (SELECT COUNT(*) FROM EMPLEADO) +
                    (SELECT COUNT(*) FROM RESERVA) +
                    (SELECT COUNT(*) FROM HABITACION) +
                    (SELECT COUNT(*) FROM USUARIO) +
                    (SELECT COUNT(*) FROM TURNO) +
                    (SELECT COUNT(*) FROM DETALLE_TURNO) as total
            """)
            total_registros = cursor.fetchone()[0]
            
            # Información de la BD
            # Obtener nombre de la BD
            cursor.execute("SELECT DATABASE()")
            db_name_info = cursor.fetchone()
            db_name = db_name_info[0] if db_name_info and db_name_info[0] else "N/A"
            
            # Obtener versión MySQL
            cursor.execute("SELECT VERSION()")
            version_info = cursor.fetchone()
            db_version = version_info[0] if version_info and version_info[0] else "N/A"
            
            # Obtener charset
            cursor.execute("SELECT @@character_set_database")
            charset_info = cursor.fetchone()
            db_charset = charset_info[0] if charset_info and charset_info[0] else "N/A"
            
            # Obtener tamaño de la BD en MB
            cursor.execute("""
                SELECT 
                    ROUND(SUM(data_length + index_length) / 1024 / 1024, 2) AS 'size_mb'
                FROM information_schema.tables 
                WHERE table_schema = DATABASE()
            """)
            size_info = cursor.fetchone()
            db_size = size_info[0] if size_info and size_info[0] else 0
            
            # Obtener número de conexiones activas/procesos
            cursor.execute("SHOW PROCESSLIST")
            procesos = cursor.fetchall()
            db_connections = len(procesos) if procesos else 0
            
            # Obtener total de índices en la BD
            cursor.execute("""
                SELECT COUNT(*) 
                FROM information_schema.statistics 
                WHERE table_schema = DATABASE()
            """)
            indices_info = cursor.fetchone()
            db_indices = indices_info[0] if indices_info and indices_info[0] else 0
            
            # Habitaciones por categoría
            cursor.execute("""
                SELECT 
                    c.nombre_categoria,
                    COUNT(*) as cantidad
                FROM HABITACION h
                INNER JOIN CATEGORIA c ON h.id_categoria = c.categoria_id
                GROUP BY c.categoria_id, c.nombre_categoria
            """)
            resultados_habitaciones = cursor.fetchall()
            habitaciones_por_categoria = {}
            for row in resultados_habitaciones:
                habitaciones_por_categoria[row[0]] = row[1]
            
            # Reservas por mes
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(fecha_registro, '%Y-%m') as mes,
                    COUNT(*) as cantidad
                FROM RESERVA
                WHERE fecha_registro >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
                GROUP BY DATE_FORMAT(fecha_registro, '%Y-%m')
                ORDER BY mes
            """)
            resultados_meses = cursor.fetchall()
            reservas_por_mes = {}
            meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
            for row in resultados_meses:
                fecha_str = row[0]
                año, mes = fecha_str.split('-')
                mes_nombre = f"{meses[int(mes)-1]} {año}"
                reservas_por_mes[mes_nombre] = row[1]
            
            # Estadísticas financieras (todas las reservas con monto_total > 0)
            cursor.execute("SELECT SUM(monto_total) FROM RESERVA WHERE monto_total > 0")
            total_ingresos = cursor.fetchone()[0] or 0
            
            cursor.execute("SELECT AVG(monto_total) FROM RESERVA WHERE monto_total > 0")
            promedio_reserva = cursor.fetchone()[0] or 0
            
            cursor.execute("SELECT MAX(monto_total) FROM RESERVA WHERE monto_total > 0")
            reserva_maxima = cursor.fetchone()[0] or 0
            
            cursor.execute("SELECT MIN(monto_total) FROM RESERVA WHERE monto_total > 0")
            reserva_minima = cursor.fetchone()[0] or 0
            
            # Ingresos por mes (últimos 12 meses, todas las reservas con monto > 0)
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(fecha_registro, '%Y-%m') as mes,
                    SUM(monto_total) as total_ingresos
                FROM RESERVA
                WHERE fecha_registro >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
                  AND monto_total > 0
                GROUP BY DATE_FORMAT(fecha_registro, '%Y-%m')
                ORDER BY mes
            """)
            resultados_ingresos_mes = cursor.fetchall()
            ingresos_por_mes = {}
            for row in resultados_ingresos_mes:
                fecha_str = row[0]
                año, mes = fecha_str.split('-')
                mes_nombre = f"{meses[int(mes)-1]} {año}"
                ingresos_por_mes[mes_nombre] = float(row[1]) if row[1] else 0
            
            return {
                'total_clientes': total_clientes,
                'total_habitaciones': total_habitaciones,
                'total_reservas': total_reservas,
                'reservas_por_estado': reservas_por_estado,
                'clientes_por_tipo': clientes_por_tipo,
                'empleados_por_turno': empleados_por_turno,
                'total_tablas': total_tablas,
                'total_registros': total_registros,
                'habitaciones_por_categoria': habitaciones_por_categoria,
                'reservas_por_mes': reservas_por_mes,
                'total_ingresos': float(total_ingresos),
                'promedio_reserva': float(promedio_reserva),
                'reserva_maxima': float(reserva_maxima),
                'reserva_minima': float(reserva_minima),
                'ingresos_por_mes': ingresos_por_mes,
                'db_name': db_name,
                'db_version': db_version,
                'db_charset': db_charset,
                'db_size': float(db_size) if db_size else 0,
                'db_connections': db_connections,
                'db_indices': db_indices
            }
    finally:
        connection.close()

def get_total_clientes():
    """
    Obtener el total de clientes
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM CLIENTE")
            total = cursor.fetchone()[0]
            return total
    finally:
        connection.close()

def get_total_habitaciones_disponibles():
    """
    Obtener el total de habitaciones disponibles (estado = 1)
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM HABITACION WHERE estado = 1")
            total = cursor.fetchone()[0]
            return total
    finally:
        connection.close()

def get_total_reservas():
    """
    Obtener el total de reservas
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM RESERVA")
            total = cursor.fetchone()[0]
            return total
    finally:
        connection.close()

def get_reservas_por_estado():
    """
    Obtener el conteo de reservas por estado
    Retorna un diccionario con {estado: cantidad}
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    CASE 
                        WHEN estado = 1 THEN 'Activa'
                        WHEN estado = 0 THEN 'Cancelada'
                        ELSE 'Desconocida'
                    END as estado_reserva,
                    COUNT(*) as cantidad
                FROM RESERVA
                GROUP BY estado
            """)
            resultados = cursor.fetchall()
            
            # Convertir a diccionario
            reservas_por_estado = {}
            for row in resultados:
                reservas_por_estado[row[0]] = row[1]
            
            return reservas_por_estado
    finally:
        connection.close()

def get_clientes_por_tipo():
    """
    Obtener el conteo de clientes por tipo
    Retorna un diccionario con {tipo: cantidad}
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    tc.descripcion as tipo_cliente,
                    COUNT(*) as cantidad
                FROM CLIENTE c
                INNER JOIN TIPO_CLIENTE tc ON c.id_tipo_cliente = tc.tipo_cliente_id
                GROUP BY tc.descripcion
            """)
            resultados = cursor.fetchall()
            
            # Convertir a diccionario
            clientes_por_tipo = {}
            for row in resultados:
                clientes_por_tipo[row[0]] = row[1]
            
            return clientes_por_tipo
    finally:
        connection.close()

def get_empleados_por_turno():
    """
    Obtener el conteo de empleados asignados por turno
    Retorna un diccionario con {turno: cantidad}
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    t.nombre_turno,
                    COUNT(DISTINCT dt.empleado_id) as cantidad
                FROM TURNO t
                LEFT JOIN DETALLE_TURNO dt ON t.turno_id = dt.turno_id
                GROUP BY t.turno_id, t.nombre_turno
            """)
            resultados = cursor.fetchall()
            
            # Convertir a diccionario
            empleados_por_turno = {}
            for row in resultados:
                empleados_por_turno[row[0]] = row[1]
            
            return empleados_por_turno
    finally:
        connection.close()

def get_total_tablas():
    """
    Obtener el total de tablas en la base de datos
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT COUNT(*) 
                FROM information_schema.tables 
                WHERE table_schema = DATABASE()
            """)
            total = cursor.fetchone()[0]
            return total
    finally:
        connection.close()

def get_total_registros():
    """
    Obtener el total de registros en todas las tablas principales
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            # Sumar registros de las tablas principales
            cursor.execute("""
                SELECT 
                    (SELECT COUNT(*) FROM CLIENTE) +
                    (SELECT COUNT(*) FROM EMPLEADO) +
                    (SELECT COUNT(*) FROM RESERVA) +
                    (SELECT COUNT(*) FROM HABITACION) +
                    (SELECT COUNT(*) FROM USUARIO) +
                    (SELECT COUNT(*) FROM TURNO) +
                    (SELECT COUNT(*) FROM DETALLE_TURNO) as total
            """)
            total = cursor.fetchone()[0]
            return total
    finally:
        connection.close()

def get_habitaciones_por_categoria():
    """
    Obtener el conteo de habitaciones por categoría
    Retorna un diccionario con {categoria: cantidad}
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    c.nombre_categoria,
                    COUNT(*) as cantidad
                FROM HABITACION h
                INNER JOIN CATEGORIA c ON h.id_categoria = c.categoria_id
                GROUP BY c.categoria_id, c.nombre_categoria
            """)
            resultados = cursor.fetchall()
            
            # Convertir a diccionario
            habitaciones_por_categoria = {}
            for row in resultados:
                habitaciones_por_categoria[row[0]] = row[1]
            
            return habitaciones_por_categoria
    finally:
        connection.close()

def get_reservas_por_mes():
    """
    Obtener el conteo de reservas por mes (últimos 12 meses)
    Retorna un diccionario con {mes: cantidad}
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(fecha_registro, '%Y-%m') as mes,
                    COUNT(*) as cantidad
                FROM RESERVA
                WHERE fecha_registro >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
                GROUP BY DATE_FORMAT(fecha_registro, '%Y-%m')
                ORDER BY mes
            """)
            resultados = cursor.fetchall()
            
            # Convertir a diccionario con formato de mes legible
            reservas_por_mes = {}
            meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
            
            for row in resultados:
                # Formato: YYYY-MM
                fecha_str = row[0]
                año, mes = fecha_str.split('-')
                mes_nombre = f"{meses[int(mes)-1]} {año}"
                reservas_por_mes[mes_nombre] = row[1]
            
            return reservas_por_mes
    finally:
        connection.close()


def generar_reporte_pdf(estadisticas, empleados_data, usuario_nombre='Usuario Desconocido'):
    """
    Generar un reporte en formato PDF con todas las estadísticas y gráficos
    Retorna el contenido del PDF como bytes
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    from datetime import datetime
    import matplotlib
    matplotlib.use('Agg')  # Backend sin GUI
    import matplotlib.pyplot as plt
    import numpy as np
    
    # Crear buffer en memoria
    buffer = BytesIO()
    
    # Crear documento PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           rightMargin=30, leftMargin=30,
                           topMargin=30, bottomMargin=30)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#328336'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        spaceBefore=20
    )
    
    # Contenido del documento
    story = []
    
    # Función auxiliar para crear gráficos
    def crear_grafico_barras(datos, titulo, ancho=6, alto=4):
        """Crea un gráfico de barras y lo retorna como BytesIO"""
        fig, ax = plt.subplots(figsize=(ancho, alto))
        labels = list(datos.keys())
        values = list(datos.values())
        colors_list = ['#1976d2', '#7b1fa2', '#e65100', '#388e3c', '#dc3545', '#ffc107']
        ax.bar(labels, values, color=colors_list[:len(labels)])
        ax.set_title(titulo, fontsize=14, fontweight='bold', pad=15)
        ax.set_ylabel('Cantidad', fontsize=10)
        ax.set_xlabel('Categoría', fontsize=10)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        return img_buffer
    
    def crear_grafico_pastel(datos, titulo, ancho=6, alto=4):
        """Crea un gráfico de pastel y lo retorna como BytesIO"""
        fig, ax = plt.subplots(figsize=(ancho, alto))
        labels = list(datos.keys())
        values = list(datos.values())
        colors_list = ['#388e3c', '#dc3545', '#ffc107', '#1976d2', '#7b1fa2', '#e65100']
        ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90, 
               colors=colors_list[:len(labels)], textprops={'fontsize': 9})
        ax.set_title(titulo, fontsize=14, fontweight='bold', pad=15)
        plt.tight_layout()
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        return img_buffer
    
    def crear_grafico_linea(datos, titulo, ancho=8, alto=4):
        """Crea un gráfico de línea y lo retorna como BytesIO"""
        fig, ax = plt.subplots(figsize=(ancho, alto))
        labels = list(datos.keys())
        values = list(datos.values())
        ax.plot(labels, values, marker='o', linewidth=2, markersize=6, color='#388e3c')
        ax.fill_between(labels, values, alpha=0.3, color='#388e3c')
        ax.set_title(titulo, fontsize=14, fontweight='bold', pad=15)
        ax.set_ylabel('Cantidad', fontsize=10)
        ax.set_xlabel('Mes', fontsize=10)
        plt.xticks(rotation=45, ha='right')
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        return img_buffer
    
    def crear_grafico_linea_dinero(datos, titulo, ancho=8, alto=4):
        """Crea un gráfico de línea para datos monetarios"""
        fig, ax = plt.subplots(figsize=(ancho, alto))
        labels = list(datos.keys())
        values = list(datos.values())
        ax.plot(labels, values, marker='o', linewidth=2, markersize=6, color='#FF6B35')
        ax.fill_between(labels, values, alpha=0.3, color='#FF6B35')
        ax.set_title(titulo, fontsize=14, fontweight='bold', pad=15)
        ax.set_ylabel('Ingresos (S/)', fontsize=10)
        ax.set_xlabel('Mes', fontsize=10)
        plt.xticks(rotation=45, ha='right')
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        return img_buffer
    
    # ========== HOJA 1: Resumen General + Clientes por Tipo ==========
    story.append(Paragraph("HOSTAL BOLÍVAR", title_style))
    story.append(Paragraph("Reporte de Estadísticas y Análisis", styles['Heading2']))
    story.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal']))
    story.append(Paragraph(f"Generado por: {usuario_nombre}", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Estadísticas principales
    story.append(Paragraph("Resumen General", heading_style))
    data_resumen = [
        ['Métrica', 'Valor'],
        ['Total Clientes', str(estadisticas.get('total_clientes', 0))],
        ['Habitaciones Disponibles', str(estadisticas.get('total_habitaciones', 0))],
        ['Total Reservas', str(estadisticas.get('total_reservas', 0))],
        ['Total Tablas en BD', str(estadisticas.get('total_tablas', 0))],
        ['Total Registros', str(estadisticas.get('total_registros', 0))]
    ]
    
    tabla_resumen = Table(data_resumen, colWidths=[4*inch, 2*inch])
    tabla_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#328336')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    story.append(tabla_resumen)
    story.append(Spacer(1, 0.3*inch))
    
    # Clientes por tipo
    if estadisticas.get('clientes_por_tipo'):
        story.append(Paragraph("CLIENTES POR TIPO", heading_style))
        # Gráfico
        img_buffer = crear_grafico_pastel(estadisticas['clientes_por_tipo'], 'Distribución de Clientes por Tipo')
        img = Image(img_buffer, width=5*inch, height=3*inch)
        story.append(img)
        story.append(Spacer(1, 0.2*inch))
        
        # Tabla de datos
        data_clientes = [['Tipo de Cliente', 'Cantidad']]
        for tipo, cantidad in estadisticas['clientes_por_tipo'].items():
            data_clientes.append([tipo, str(cantidad)])
        
        tabla_clientes = Table(data_clientes, colWidths=[4*inch, 2*inch])
        tabla_clientes.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7b1fa2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        story.append(tabla_clientes)
    
    story.append(PageBreak())
    
    # ========== HOJA 2: Reservas por Mes + Ingresos por Mes + Análisis Financiero (vertical, compacto) ==========
    # Reservas por mes
    if estadisticas.get('reservas_por_mes'):
        story.append(Paragraph("RESERVAS POR MES (ÚLTIMOS 12 MESES)", heading_style))
        # Gráfico más pequeño
        img_buffer = crear_grafico_linea(estadisticas['reservas_por_mes'], 'Tendencia de Reservas por Mes')
        img = Image(img_buffer, width=5.5*inch, height=2.5*inch)
        story.append(img)
        story.append(Spacer(1, 0.1*inch))
        
        # Tabla de datos más compacta
        data_meses = [['Mes', 'Cantidad']]
        for mes, cantidad in estadisticas['reservas_por_mes'].items():
            data_meses.append([mes, str(cantidad)])
        
        tabla_meses = Table(data_meses, colWidths=[3*inch, 1.2*inch])
        tabla_meses.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#328336')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        story.append(tabla_meses)
        story.append(Spacer(1, 0.15*inch))
    
    # Ingresos por mes
    if estadisticas.get('ingresos_por_mes') and len(estadisticas['ingresos_por_mes']) > 0:
        story.append(Paragraph("INGRESOS POR MES (ÚLTIMOS 12 MESES)", heading_style))
        # Gráfico más pequeño
        img_buffer = crear_grafico_linea_dinero(estadisticas['ingresos_por_mes'], 'Tendencia de Ingresos por Mes')
        img = Image(img_buffer, width=5.5*inch, height=2.5*inch)
        story.append(img)
        story.append(Spacer(1, 0.1*inch))
        
        # Tabla de datos más compacta
        data_ingresos = [['Mes', 'Ingresos (S/)']]
        for mes, ingreso in estadisticas['ingresos_por_mes'].items():
            data_ingresos.append([mes, f"S/ {ingreso:,.2f}"])
        
        tabla_ingresos = Table(data_ingresos, colWidths=[3*inch, 1.2*inch])
        tabla_ingresos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B35')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        story.append(tabla_ingresos)
        story.append(Spacer(1, 0.15*inch))
    
    # Análisis Financiero al final
    story.append(Paragraph("ANÁLISIS FINANCIERO", heading_style))
    
    # Resumen financiero más compacto
    data_financiero = [
        ['Métrica Financiera', 'Valor'],
        ['Total Ingresos (S/)', f"S/ {estadisticas.get('total_ingresos', 0):,.2f}"],
        ['Promedio por Reserva (S/)', f"S/ {estadisticas.get('promedio_reserva', 0):,.2f}"],
        ['Reserva Máxima (S/)', f"S/ {estadisticas.get('reserva_maxima', 0):,.2f}"],
        ['Reserva Mínima (S/)', f"S/ {estadisticas.get('reserva_minima', 0):,.2f}"]
    ]
    
    tabla_financiero = Table(data_financiero, colWidths=[3*inch, 1.2*inch])
    tabla_financiero.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B35')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    story.append(tabla_financiero)
    
    # Construir PDF
    doc.build(story)
    
    # Obtener el contenido del buffer
    buffer.seek(0)
    return buffer.read()


def generar_reporte_excel(estadisticas, empleados_data, usuario_nombre='Usuario Desconocido'):
    """
    Generar un reporte en formato Excel con todas las estadísticas
    Retorna el contenido del Excel como bytes
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from datetime import datetime
    except ImportError:
        raise ImportError("openpyxl no está instalado. Por favor ejecuta: pip install openpyxl")
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Estadísticas"
    
    # Estilos
    header_fill = PatternFill(start_color="328336", end_color="328336", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Título
    ws.merge_cells('A1:B1')
    ws['A1'] = "HOSTAL BOLÍVAR - Reporte de Estadísticas"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    
    ws.merge_cells('A2:B2')
    ws['A2'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws['A2'].alignment = center_alignment
    
    ws.merge_cells('A3:B3')
    ws['A3'] = f"Generado por: {usuario_nombre}"
    ws['A3'].alignment = center_alignment
    
    row = 5
    
    # Resumen general
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "RESUMEN GENERAL"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    ws['A' + str(row)] = 'Métrica'
    ws['B' + str(row)] = 'Valor'
    ws['A' + str(row)].fill = header_fill
    ws['A' + str(row)].font = header_font
    ws['B' + str(row)].fill = header_fill
    ws['B' + str(row)].font = header_font
    ws['A' + str(row)].alignment = center_alignment
    ws['B' + str(row)].alignment = center_alignment
    row += 1
    
    resumen_data = [
        ['Total Clientes', estadisticas.get('total_clientes', 0)],
        ['Habitaciones Disponibles', estadisticas.get('total_habitaciones', 0)],
        ['Total Reservas', estadisticas.get('total_reservas', 0)],
        ['Total Tablas en BD', estadisticas.get('total_tablas', 0)],
        ['Total Registros', estadisticas.get('total_registros', 0)]
    ]
    
    for item in resumen_data:
        ws['A' + str(row)] = item[0]
        ws['B' + str(row)] = item[1]
        ws['A' + str(row)].border = border
        ws['B' + str(row)].border = border
        row += 1
    
    row += 2
    
    # Clientes por tipo
    if estadisticas.get('clientes_por_tipo'):
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "CLIENTES POR TIPO"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        ws['A' + str(row)] = 'Tipo de Cliente'
        ws['B' + str(row)] = 'Cantidad'
        ws['A' + str(row)].fill = PatternFill(start_color="7b1fa2", end_color="7b1fa2", fill_type="solid")
        ws['A' + str(row)].font = header_font
        ws['B' + str(row)].fill = PatternFill(start_color="7b1fa2", end_color="7b1fa2", fill_type="solid")
        ws['B' + str(row)].font = header_font
        ws['A' + str(row)].alignment = center_alignment
        ws['B' + str(row)].alignment = center_alignment
        row += 1
        
        start_row_clientes = row
        for tipo, cantidad in estadisticas['clientes_por_tipo'].items():
            ws['A' + str(row)] = tipo
            ws['B' + str(row)] = cantidad
            ws['A' + str(row)].border = border
            ws['B' + str(row)].border = border
            row += 1
        
        row += 2
    
    # Reservas por mes
    if estadisticas.get('reservas_por_mes'):
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "RESERVAS POR MES (ÚLTIMOS 12 MESES)"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        ws['A' + str(row)] = 'Mes'
        ws['B' + str(row)] = 'Cantidad'
        ws['A' + str(row)].fill = header_fill
        ws['A' + str(row)].font = header_font
        ws['B' + str(row)].fill = header_fill
        ws['B' + str(row)].font = header_font
        ws['A' + str(row)].alignment = center_alignment
        ws['B' + str(row)].alignment = center_alignment
        row += 1
        
        for mes, cantidad in estadisticas['reservas_por_mes'].items():
            ws['A' + str(row)] = mes
            ws['B' + str(row)] = cantidad
            ws['A' + str(row)].border = border
            ws['B' + str(row)].border = border
            row += 1
        
        row += 2
    
    # Estadísticas de dinero
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "ANÁLISIS FINANCIERO"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    ws['A' + str(row)] = 'Métrica Financiera'
    ws['B' + str(row)] = 'Valor'
    ws['A' + str(row)].fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    ws['A' + str(row)].font = header_font
    ws['B' + str(row)].fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    ws['B' + str(row)].font = header_font
    ws['A' + str(row)].alignment = center_alignment
    ws['B' + str(row)].alignment = center_alignment
    row += 1
    
    financiero_data = [
        ['Total Ingresos (S/)', f"S/ {estadisticas.get('total_ingresos', 0):,.2f}"],
        ['Promedio por Reserva (S/)', f"S/ {estadisticas.get('promedio_reserva', 0):,.2f}"],
        ['Reserva Máxima (S/)', f"S/ {estadisticas.get('reserva_maxima', 0):,.2f}"],
        ['Reserva Mínima (S/)', f"S/ {estadisticas.get('reserva_minima', 0):,.2f}"]
    ]
    
    for item in financiero_data:
        ws['A' + str(row)] = item[0]
        ws['B' + str(row)] = item[1]
        ws['A' + str(row)].border = border
        ws['B' + str(row)].border = border
        row += 1
    
    row += 2
    
    # Ingresos por mes
    if estadisticas.get('ingresos_por_mes'):
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "INGRESOS POR MES (ÚLTIMOS 12 MESES)"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        ws['A' + str(row)] = 'Mes'
        ws['B' + str(row)] = 'Ingresos (S/)'
        ws['A' + str(row)].fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        ws['A' + str(row)].font = header_font
        ws['B' + str(row)].fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        ws['B' + str(row)].font = header_font
        ws['A' + str(row)].alignment = center_alignment
        ws['B' + str(row)].alignment = center_alignment
        row += 1
        
        for mes, ingreso in estadisticas['ingresos_por_mes'].items():
            ws['A' + str(row)] = mes
            ws['B' + str(row)] = f"S/ {ingreso:,.2f}"
            ws['A' + str(row)].border = border
            ws['B' + str(row)].border = border
            row += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    
    # Reducir el espaciado entre filas y ajustar alineación
    for r in range(1, row + 1):
        if ws[f'A{r}'].value is not None or ws[f'B{r}'].value is not None:
            ws.row_dimensions[r].height = 20
            if ws[f'A{r}'].value is not None:
                ws[f'A{r}'].alignment = Alignment(vertical='center', horizontal='left')
            if ws[f'B{r}'].value is not None:
                ws[f'B{r}'].alignment = Alignment(vertical='center', horizontal='center')
    
    # Guardar en buffer
    buffer = BytesIO()
    try:
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise Exception(f"Error al guardar Excel: {str(e)}")

def get_lista_tablas():
    """
    Obtener la lista de todas las tablas en la base de datos
    Retorna una lista de nombres de tablas ordenadas alfabéticamente
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT table_name 
                FROM information_schema.tables 
                WHERE table_schema = DATABASE()
                ORDER BY table_name
            """)
            tablas = cursor.fetchall()
            return [tabla[0] for tabla in tablas]
    finally:
        connection.close()

def get_atributos_tabla(nombre_tabla):
    """
    Obtener los atributos (columnas) de una tabla específica
    Retorna una lista de diccionarios con información de cada columna
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    COLUMN_NAME as nombre,
                    DATA_TYPE as tipo_dato,
                    CHARACTER_MAXIMUM_LENGTH as longitud_maxima,
                    IS_NULLABLE as permite_nulo,
                    COLUMN_KEY as clave,
                    COLUMN_DEFAULT as valor_default,
                    EXTRA as extra
                FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE()
                AND TABLE_NAME = %s
                ORDER BY ORDINAL_POSITION
            """, (nombre_tabla,))
            columnas = cursor.fetchall()
            resultado = []
            for col in columnas:
                atributo = {
                    'nombre': col[0],
                    'tipo_dato': col[1],
                    'longitud_maxima': col[2] if col[2] else None,
                    'permite_nulo': col[3],
                    'clave': col[4] if col[4] else '',
                    'valor_default': col[5] if col[5] else None,
                    'extra': col[6] if col[6] else ''
                }
                resultado.append(atributo)
            return resultado
    finally:
        connection.close()

def get_detalle_conexiones():
    """
    Obtener el detalle de todas las conexiones activas (SHOW PROCESSLIST)
    Retorna una lista de diccionarios con la información de cada proceso
    """
    import pymysql.cursors
    connection = get_connection()
    try:
        with connection.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SHOW PROCESSLIST")
            procesos = cursor.fetchall()
            
            # Convertir a lista de diccionarios
            resultado = []
            for proceso in procesos:
                proceso_dict = {}
                for key, value in proceso.items():
                    proceso_dict[key] = value
                resultado.append(proceso_dict)
            
            return resultado
    finally:
        connection.close()

